import openpyxl
import unittest
from excel_project import file_1


class deg_reg_test(unittest.TestCase):
    file1='Source_file.xlsx'
    file2='date_hour.xlsx'


    def test_number_cell(self):
        wb=openpyxl.load_workbook(self.file1)
        sheet=wb.active
        no_value=sheet['A2'].value
        self.assertIsNotNone(no_value)

    def test_degree_cell(self):
        wb=openpyxl.load_workbook(self.file1)
        sheet=wb.active
        degree_value=sheet['B2'].value
        self.assertIsNotNone(degree_value)

    def test_reset_data_source_file(self):
        wb = openpyxl.load_workbook(self.file1)
        sheet = wb.active
        no_value=sheet['A2'].value
        degree_value = sheet['B2'].value
        self.assertIsNone(degree_value,no_value)

    def test_reset_data_date_hour(self):
        wb = openpyxl.load_workbook(self.file2)
        sheet = wb.active
        cells=['A2','B2','C2','D2']
        for i in cells:
            self.assertIsNone(sheet[i].value)

    # def test_askopenfile1(self):
    #     file_1(0)
    #     expected='Source_file.xlsx'
    #     actual=FILE1
    #     self.assertEqual(actual,expected)