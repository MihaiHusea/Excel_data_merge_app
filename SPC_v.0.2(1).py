"""
SPC_v0.2
 sheet['D8'] = diagramaControl
"""
print('SPC_v.2___Aug/2022 ___ Rumagol')

import datetime
import xlrd
from openpyxl import load_workbook
import tkinter as tk
from tkinter.filedialog import askopenfilename

file1 = ''
file2 = ''
file3 = 'date_recorder.xlsx'
# now = datetime.now()
# year = now.strftime("%Y")

# pentru ca nu functioneaza variabilele astea globale, file1, file2
ar = ['', '']


def operatie(event):
    print(ar)
    file1 = ar[0]
    file2 = ar[1]
    # file3 = ar[2]

    print("File names received.")
    print(file1)
    print(file2)
    print(file3)
    # To open Workbook
    print("Opening " + file1)
    wb = xlrd.open_workbook(file1)
    sheet = wb.sheet_by_index(0)
    sheet.cell_value(0, 0)

    print("Reading values from " + file1 + "...")
    upperLimit = sheet.cell_value(12, 3)  # e13
    lowerLimit = sheet.cell_value(13, 3)  # e14
    referinta = sheet.cell_value(5, 3)  # e16
    # diagramaControl = sheet.cell_value(6, 3)  # e17
    cp = sheet.cell_value(23, 3)  # e25
    cpk = sheet.cell_value(24, 3)  # e26
    machine = sheet.cell_value(7, 19)  # t8
    dataa = sheet.cell_value(5, 19)  # t8
    caracteristica = sheet.cell_value(9, 3)

    # Begin process to write into result file
    print("Opening " + file1 + "...")
    wb = load_workbook(file1)
    sheet = wb.active

    sheet['A60'] = 30

    print("Saving file " + file1)
    wb.save(file1)
    # print("Done!")

    print("Opening " + file1)
    wb = xlrd.open_workbook(file1)
    sheet = wb.sheet_by_index(0)
    sheet.cell_value(0, 0)

    # aggregating values from vertical to list
    componentValues = []

    for i in range(30, 60):  # eroare
        componentValues.append(sheet.cell_value(i, 3))

    # data , ora import
    print("Opening " + file3)
    wb = xlrd.open_workbook(file3)
    sheet = wb.sheet_by_index(0)
    sheet.cell_value(0, 0)

    # data
    data_value = []
    ora_value = []

    for i in range(1,30):
        data_value.append(sheet.cell_value(i,1))
        ora_value.append(sheet.cell_value(i,2))

    print("Opening " + file2 + "...")
    wb = load_workbook(file2)
    sheet = wb.active

    # mutarea valorilor de pe cele 30 de pozitii pe verticala
    # in celalalt sheet pe orizontala
    print("Writing values in " + file2)
    coloana = 4
    for nr in componentValues:
        c1 = sheet.cell(row=23, column=coloana)
        c1.value = nr
        coloana = coloana + 1

    coloana = 4
    for d in data_value:
        d1 = sheet.cell(row=21, column=coloana)
        d1.value = d
        coloana += 1
    coloana = 4
    for t in ora_value:
        t1 = sheet.cell(row=22, column=coloana)
        t1.value = t
        coloana += 1

    # year
    year_time = datetime.datetime.now()
    year_1 = year_time.year

    sheet['P17'] =year_1
    sheet['D13'] = upperLimit
    sheet['D14'] = lowerLimit
    sheet['J13'] = cp
    sheet['J14'] = cpk
    sheet['D10'] = machine
    # sheet['D8'] = diagramaControl
    sheet['AD4'] = dataa
    sheet['C12'] = caracteristica

    sheet.merge_cells('AA10:AG10')
    sheet['AA10'] = referinta

    print("Saving file " + file2)
    wb.save(file2)
    print("Done!")


window = tk.Tk()
window.geometry("200x350")
butonFisier1 = tk.Button(
    text="Fisier sursa",
    width=15,
    height=3,
    font=("Arial", 15, "bold"),
    bg="green",
    fg="black",
)
butonFisier1.pack()

butonFisier2 = tk.Button(
    text="Fisier destinatie",
    width=15,
    height=3,
    font=("Arial", 15, "bold"),
    bg="green",
    fg="black",
)
butonFisier2.pack()

butonOperatie = tk.Button(
    text="Executie!",
    width=15,
    height=2,
    font=("Arial", 15, "bold"),
    bg="green",
    fg="black",
)
butonOperatie.pack()

buton_resetare_data = tk.Button(
    text="Resetare Data",
    width=15,
    height=3,
    font=("Arial", 15, "bold"),
    bg="red",
    fg="black",
)
buton_resetare_data.pack()


def fisier1(event):
    file1 = askopenfilename(
        filetypes=[("Text Files", "*.xlsx"), ("All Files", "*.*")]
    )
    ar[0] = file1


def fisier2(event):
    file2 = askopenfilename(
        filetypes=[("Text Files", "*.xlsx"), ("All Files", "*.*")]
    )
    ar[1] = file2


# deleting
def date_reset(event):
    file3 = "date_recorder.xlsx"
    wb = load_workbook(file3)
    sheet = wb.active
    # sterge coloane
    sheet.delete_cols(5)
    sheet.delete_cols(4)
    sheet.delete_cols(3)
    sheet.delete_cols(2)

    # scriere
    data_str = sheet.cell(row=1, column=2)
    ora_str = sheet.cell(row=1, column=3)
    data_now_str = sheet.cell(row=1, column=4)
    data_time_sec = sheet.cell(row=2, column=4)

    d1 = 'data'
    d2 = 'time_time'
    h1 = 'ora'
    d3 = 0

    data_str.value = d1
    data_now_str.value = d2
    ora_str.value = h1
    data_time_sec.value = d3

    wb.save(file3)
    print('deleting ...' + file3)
    print('saving...' + file3)


butonFisier1.bind("<Button-1>", fisier1)
butonFisier2.bind("<Button-1>", fisier2)
butonOperatie.bind("<Button-1>", operatie)
buton_resetare_data.bind("<Button-1>", date_reset)

window.mainloop()
