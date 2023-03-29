#!/usr/bin/env python3
"""
Excel automation project
"""
import time
import tkinter as tk
import random
from tkinter.filedialog import askopenfilename
from datetime import datetime
import openpyxl
from tkinter import *
import os
import sys
import subprocess


def select_source_file1():
    global file1
    data_load_confirmation_label['fg'] = '#1D6F42'
    file1 = askopenfilename(
        filetypes=[('file1', 'measurements.xlsx'), ('all files', '*.*')])
    data_load_confirmation_label['fg'] = 'black'
    if file1:
        data_load_confirmation_label['text'] = str(file1) + ' loaded!'
    else:
        data_load_confirmation_label['fg'] = 'white'
        data_load_confirmation_label[
            'text'] = 'No file selected! Click "Load file(measurements)" button to select a file.'


def select_source_file_2():
    global file3
    tr_file_load_confirmation_label['fg'] = '#1D6F42'
    file3 = askopenfilename(
        filetypes=[('file3', 'date.xlsx'), ('all files', '*.*')])
    tr_file_load_confirmation_label['fg'] = 'black'
    if file3:
        tr_file_load_confirmation_label['text'] = str(file3) + ' loaded!'
    else:
        tr_file_load_confirmation_label['fg'] = 'white'
        tr_file_load_confirmation_label['text'] = 'No file selected! Click "Load file(date)" button to select a file.'


def select_report_file():
    global file2
    report_file_load_confirmation_label['fg'] = '#1D6F42'
    file2 = askopenfilename(
        filetypes=[('file2', 'report.xlsx'), ('all files', '*.*')])
    report_file_load_confirmation_label['fg'] = 'black'
    if file2:
        report_file_load_confirmation_label['text'] = str(file2) + ' loaded !'
    else:
        report_file_load_confirmation_label['fg'] = 'white'
        report_file_load_confirmation_label[
            'text'] = 'No file selected! Click "Load file(report)" button to select a file.'


def record_data():
    try:
        record_data_confirmation_label['fg'] = 'black'
        # write date and hour:
        wb = openpyxl.load_workbook(file3)
        sheet = wb['Sheet']

        date_now = datetime.now()  # current date and hour
        hour_now = date_now.strftime('%H:%M:%S')  # hour format
        day_now = date_now.strftime('%d''-''%m''-''%y')  # day format
        epoch = time.time()  # epoch time in seconds (from 01.01.1970)

        number_cell = ['A' + str(i) for i in range(1, 22)]
        hour_cell = ['B' + str(i) for i in range(1, 22)]
        date_cell = ['C' + str(i) for i in range(1, 22)]
        epoch_cell = ['D' + str(i) for i in range(1, 22)]

        count = 1
        delta = 6
        wait_sec = 5

        for i in epoch_cell[1:21]:
            if sheet[i].value is not None:
                delta = epoch - sheet[i].value
                count += 1
                if count == 21:
                    record_data_confirmation_label['text'] = 'Full memory! Press delete for reset data!'
            elif delta > wait_sec:
                sheet[i].value = epoch
                sheet[date_cell[count]].value = day_now
                sheet[hour_cell[count]].value = hour_now
                sheet[number_cell[count]].value = str(count) + '.'
                record_data_confirmation_label['text'] = 'Values have been recorded!'
                record_temperature()  # write temperature data
                break
            else:
                record_data_confirmation_label[
                    'text'] = f'Please wait {wait_sec - int(delta)} seconds until the next record! '
                break
        wb.save(file3)
    except NameError:
        record_data_confirmation_label['fg'] = "yellow"
        record_data_confirmation_label['text'] = "Please load files first!"


def record_temperature():
    try:
        global file1
        wb = openpyxl.load_workbook(file1)
        sheet = wb['Sheet']
        degree = random.randint(18, 22)
        number_cell = ['A' + str(i) for i in range(1, 22)]
        degree_cell = ['B' + str(i) for i in range(1, 22)]
        count = 0
        for i in number_cell[1:21]:
            count += 1
            if sheet[i].value is None:
                sheet[i].value = str(count) + '.'
                sheet[degree_cell[count]].value = int(degree)
                break
        wb.save(file1)
    except NameError:
        record_data_confirmation_label['fg'] = "yellow"
        record_data_confirmation_label['text'] = "Please load files first!"


def range_letter(start, stop):
    return (chr(n) for n in range(ord(start), ord(stop) + 1))


def execute():
    try:
        execution_confirmation_label['fg'] = 'black'
        # load and read from file 1
        wb = openpyxl.load_workbook(file1)
        sheet = wb.active
        no_list = []
        t_list = []

        for i in range(1, 22):
            no_cell = sheet['A' + str(i)].value
            no_list.append(no_cell)
            temp_cell = sheet['B' + str(i)].value
            t_list.append(temp_cell)

        # load and read from FILE3
        wb = openpyxl.load_workbook(file3)
        sheet = wb.active
        h_list = []
        d_list = []
        for i in range(1, 22):
            hour_cell = sheet['B' + str(i)].value
            h_list.append(hour_cell)
            date_cell = sheet['C' + str(i)].value
            d_list.append(date_cell)

        # write report
        line_no = [str(i) + '2' for i in range_letter("A", "Z")]
        line_temp = [str(i) + '3' for i in range_letter("A", "Z")]
        line_hour = [str(i) + '4' for i in range_letter("A", "Z")]
        line_date = [str(i) + '5' for i in range_letter("A", "Z")]
        line_nominal = [str(i) + '8' for i in range_letter("A", "Z")]
        line_l_tol = [str(i) + '9' for i in range_letter("A", "Z")]
        line_u_tol = [str(i) + '10' for i in range_letter("A", "Z")]

        wb = openpyxl.load_workbook(file2)
        sheet = wb.active

        for i in range(21):
            # start_no_list= start index for each list
            start_no_list = no_list[i]
            start_t_list = t_list[i]
            start_h_list = h_list[i]
            start_d_list = d_list[i]

            sheet[line_no[i]].value = start_no_list
            sheet[line_temp[i]].value = start_t_list
            sheet[line_hour[i]].value = start_h_list
            sheet[line_date[i]].value = start_d_list

        for i in range(1, 21):
            if sheet[line_no[i]].value is not None:
                sheet[line_nominal[i]].value = 20
                sheet[line_u_tol[i]].value = 22
                sheet[line_l_tol[i]].value = 18
            else:
                sheet[line_nominal[i]].value = None
                sheet[line_u_tol[i]].value = None
                sheet[line_l_tol[i]].value = None
        execution_confirmation_label['text'] = f'Report created: {file2}'
        wb.save(file2)
    except NameError:
        execution_confirmation_label['fg'] = "yellow"
        execution_confirmation_label['text'] = "Please load files first!"


def reset_data():
    detele_confirmation_label['fg'] = 'black'
    # read and load FILE3
    global file3
    wb = openpyxl.load_workbook(file3)
    sheet = wb.active
    # delete columns
    for i in range(4):  # using for loop with range(4) to delete column 1 four times
        sheet.delete_cols(1)  # when deleted column 1 it's been replaced with column 2 and so on
    # write
    number = sheet.cell(row=1, column=1)
    hour = sheet.cell(row=1, column=2)
    date = sheet.cell(row=1, column=3)
    time = sheet.cell(row=1, column=4)

    number_title = 'No.'
    hour_title = 'Hour'
    date_title = 'Date'
    epoch_title = 'Epoch'

    number.value = number_title
    hour.value = hour_title
    date.value = date_title
    time.value = epoch_title
    wb.save(file3)

    global file1
    wb = openpyxl.load_workbook(file1)
    sheet = wb.active

    for i in range(2):
        sheet.delete_cols(1)
        nr = sheet.cell(row=1, column=1)
        degree = sheet.cell(row=1, column=2)
        number = 'No.'
        d1 = 'Temperature(°C)'
        nr.value = number
        degree.value = d1
    wb.save(file1)
    detele_confirmation_label['text'] = 'Data has been deleted!'


def show_report():
    try:
        if sys.platform == "win32":
            os.startfile(f'{file2}')
        else:
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.call([opener, file2])
    except NameError:
        open_report_confirmation_label['fg'] = 'yellow'
        open_report_confirmation_label['text'] = 'Please load report file first!.'


# GUI
window = Tk()
window.title('AppX v1.0')
window.geometry("1400x750")
window.configure(background="#1D6F42")
background_text = tk.Label(window,
                           text='KEEP\nCALM\nIT' + "'" + 'S JUST AN\nEXCEL\nFILE',
                           bg='#1D6F42',
                           font=("Arial", 35, "bold"),
                           fg="white")
background_text.place(x=20, y=165)

load_tr_data = tk.Button(
    text="Load file\n(date)",
    width=15,
    height=3,
    font=("Arial", 14, "bold"),
    bg="#abf77e",
    fg="black",
    highlightthickness=0,
    command=select_source_file_2
)
load_tr_data.place(x=550, y=20)
tr_file_load_confirmation_label = tk.Label(window, font=('Arial', 12, 'bold'), bg='#1D6F42', fg='white')
tr_file_load_confirmation_label.place(x=800, y=20, width=600, height=85)
tr_file_load_confirmation_label['text'] = 'No file selected! Click "Load file(date)" button to load a file.'

load_data_button = tk.Button(
    text="Load file\n(measurements)",
    width=15,
    height=3,
    font=("Arial", 14, "bold"),
    bg='#abf77e',
    fg="black",
    highlightthickness=0,
    command=select_source_file1,
)

load_data_button.place(x=450, y=120)
data_load_confirmation_label = Label(window, font=('Arial', 12, 'bold'), bg='#1D6F42', fg='white')
data_load_confirmation_label.place(x=700, y=120, width=600, height=85)
data_load_confirmation_label['text'] = 'No file selected! Click "Load file(measurements)" button to select a file.'

load_tr_data = tk.Button(
    text="Load file\n(report)",
    width=15,
    height=3,
    font=("Arial", 14, "bold"),
    bg="#abf77e",
    fg="black",
    highlightthickness=0,
    command=select_report_file
)
load_tr_data.place(x=550, y=220)
report_file_load_confirmation_label = tk.Label(window, font=('Arial', 12, 'bold'), bg='#1D6F42', fg='white')
report_file_load_confirmation_label.place(x=800, y=220, width=600, height=85)
report_file_load_confirmation_label['text'] = 'No file selected! Click "Load file(report)" button to select a file.'

record_data_button = tk.Button(
    text="Record\ndata",
    width=15,
    height=3,
    font=("Arial", 14, "bold"),
    bg="#abf77e",
    fg="black",
    highlightthickness=0,
    command=record_data
)
record_data_button.place(x=450, y=320)
record_data_confirmation_label = tk.Label(window, font=('Arial', 12, 'bold'), bg='#1D6F42', fg='white')
record_data_confirmation_label.place(x=650, y=320, width=600, height=85)
record_data_confirmation_label['text'] = 'Click "Rec" button to record data.'

execute_button = tk.Button(
    text="Execute",
    width=15,
    height=3,
    font=("Arial", 14, "bold"),
    bg="#3EB489",
    fg="black",
    highlightthickness=0,
    command=execute
)
execute_button.place(x=550, y=420)
execution_confirmation_label = tk.Label(window, font=('Arial', 12, 'bold'), bg='#1D6F42', fg='white')
execution_confirmation_label.place(x=750, y=420, width=600, height=85)
execution_confirmation_label['text'] = 'Click "Execute" button to create data report.'

delete_button = tk.Button(
    text="Delete",
    width=15,
    height=3,
    font=("Arial", 14, "bold"),
    bg="#E00201",
    fg="black",
    highlightthickness=0,
    command=reset_data
)
delete_button.place(x=450, y=520)
detele_confirmation_label = tk.Label(window, font=('Arial', 12, 'bold'), bg='#1D6F42', fg='white')
detele_confirmation_label.place(x=650, y=520, width=600, height=85)
detele_confirmation_label['text'] = 'Click "Delete" button to reset data.'

open_report_label = tk.Button(
    text="Open Report",
    width=15,
    height=3,
    font=("Arial", 14, "bold"),
    bg="yellow",
    fg="black",
    highlightthickness=0,
    command=show_report
)
open_report_label.place(x=550, y=620)
open_report_confirmation_label = tk.Label(window, font=('Arial', 12, 'bold'), bg='#1D6F42', fg='white')
open_report_confirmation_label.place(x=750, y=620, width=600, height=85)
open_report_confirmation_label['text'] = 'Click "Open Report" button to open file report.'

window.mainloop()
